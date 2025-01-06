from openpyxl import Workbook
import sys
from openpyxl import load_workbook

part = '｜'

def readLines(filename):
    fileList = []
    with open(filename) as f:
        fileList = f.readlines()
        return fileList



asmListPath = sys.argv[1]
outputPath = sys.argv[2]

textCommands=['page','text','line','cont','para','next']
def isTextCommand(component):
    for command in textCommands:
        if component == command:
            return True
    return False

fileList = readLines(asmListPath)

wb = load_workbook(filename = outputPath)
index = 0
for i in range(len(fileList)):
    file = fileList[index]
    fileName = file.split('/')[-1]
    ws1 = wb.get_sheet_by_name(fileName.replace('\n','').replace('.asm',''))
    lines = readLines(file.replace('\n',''))

    for row in ws1['A1:C2000']:
        for cell in row:
            cell.value = None

    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 28
    ws1.column_dimensions['D'].width = 10
    ws1.column_dimensions['E'].width = 28
    ws1.column_dimensions['F'].width = 16
    ws1.column_dimensions['G'].width = 28
    ws1.column_dimensions['H'].width = 10
    ws1.column_dimensions['I'].width = 28

    ws1.cell(row=1, column=1, value = file.replace('\n',''))
    id = 2
    for line in lines:
        line1 = line.replace(' \"',part).replace('\"',part).replace('\t','').replace('\n','')
        
        if len(line1) >= 1:
            if line1[-1] == ':':
                # it's an asm label
                ws1.cell(row=id, column=1, value = line1.replace('\n',''))
            else:
                
                components = line1.split(part)
                tmp = line1.split(' ')
                if len(tmp) > 0:
                    if '_' in tmp[0]:
                        components = line1.split(' ')
                        if len(components) >= 2:
                            txt = ''
                            for i in range(1,len(components)):
                                txt += components[i] + ' '
                            components = [tmp[0],txt]
                            # print(components)
                if len(components) >= 2:
                    if True or isTextCommand(components[0]):
                        ws1.cell(row=id, column=2, value = components[0])
                        ws1.cell(row=id, column=3, value = components[1])
                elif len(components) == 1:
                    ws1.cell(row=id, column=2, value = components[0])
                else:
                    ws1.cell(row=id, column=1, value = line1.replace('\n',''))
        id += 1
    ws1.cell(row=id, column=1, value = 'end')
    index += 1


wb.save(outputPath)




        





