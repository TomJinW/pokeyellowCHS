from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border

import sys
import openpyxl
import shutil
import os
import charmap

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

def addFilePaths(path):
    if not path in filePaths:
        filePaths.append(path)
        os.makedirs(os.path.dirname(tmp+path), exist_ok=True)
        shutil.copyfile(path,tmp+path)

def readTextFileLines(path):
    file = open(path,"r", encoding='utf-8')
    return file.readlines()

def removeNone(text):
    if text == None:
        return ''
    return text

tmp = 'tmp/'
# Program Start
xlsxListPath = sys.argv[1]
mode = int(sys.argv[2])
hexchar = int(sys.argv[3])
buildMode = int(sys.argv[4])
buildVer = sys.argv[5]

def getCategoryText(origCategory):
    category = 'db \"' + origCategory + '@\"'
    if hexchar == 1:
        category = origCategory + '#@'
        category = charmap.replaceText(category,charMap,buildMode)
        category = 'db ' + category
    return category

# Load Workbook
wb = load_workbook(filename = xlsxListPath)

filePaths = []

charMap = charmap.readCharMaps()

for sheet in wb._sheets:
	# Get init File Path
    filePath = sheet.cell(row=1, column=1).value
    print()
    print(bcolors.OKGREEN)
    print('Importing Dex Entries')
    print(filePath)
    print('.............')
	# Backup Original Files
    addFilePaths(filePath)
    textLines = readTextFileLines(filePath)

    header = ''
    labelCount = 0
    for line in textLines:
        text = line.replace('\n','')
        
        if len(text) > 0:
            if text[-1] == ':':
                labelCount += 1
                if labelCount > 1:
                    break
                else:
                    header += line
            else:
                header += line
        else:
            header += line
    id = 2

    body = ''
    while sheet.cell(row=id, column=1).value != 'end':
        
        label = sheet.cell(row=id, column=mode + 0).value
        
        origCategory = sheet.cell(row=id, column=mode + 1).value
        category = getCategoryText(origCategory)
        if buildVer == 'RGB' and origCategory == '虚拟':
            line1 = 'IF DEF(_BLUE)\n'
            line2 = '\t' + getCategoryText(origCategory) + '\n'
            line3 = '\tELSE\n'
            line4 = '\t' + getCategoryText('ＣＧ') + '\n'
            line5 = '\tENDC'
            category = line1 + line2 + line3 + line4 + line5
            
        height = 'db ' + str(sheet.cell(row=id, column=mode + 2).value).replace('-',',')
        weight = 'dw ' + str(sheet.cell(row=id, column=mode + 3).value).replace('-','')
        descValue = sheet.cell(row=id, column=mode + 4).value
        desc = 'text_far ' + descValue + '\n\ttext_end\n\n'
        desc += '\ttext_far ' + descValue + '2' + '\n\ttext_end\n\n'
        if not '_' in descValue:
            desc = 'db \"コメント さくせいちゅう@\"'
        id += 1

        body += label +'\n\t' + category + '\n\t' + height + '\n\t' + weight + '\n\t' + desc 

    with open(filePath, 'w', encoding='utf-8') as f:
        f.write(header + body)



print(bcolors.OKGREEN)
print('Dex Entries Import Complete')
print(filePath)
print('.............')
