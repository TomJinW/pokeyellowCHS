#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
@author: shaohd
@contact:838653940@qq.com
@file: LCS.py
@time: 2022/3/3 23:14
"""

"""
最长非连续公共字串提取
"""

import sys
import lcspy
from openpyxl.styles import Color, PatternFill, Font, Border
warningFill = PatternFill(start_color='E0F8E0F8',
                   end_color='E0F8E0F8',
                   fill_type='solid')

def removeNone(text):
    if text == None:
        return ''
    return str(text)
# class Temp:
#     L, idx1, idx2 = 0, 0, 0


# def WLCS(sent1, sent2):
#     len1 = len(sent1)
#     len2 = len(sent2)
#     if len1 == 0 or len2 == 0:
#         return ''

#     Matrix_sent = [[Temp() for j in range(0, len2)] for i in range(0, len1)]
#     # print(f"Matrix_sent:{len(Matrix_sent), len1, len2}")
#     if sent1[0] == sent2[0]:
#         Matrix_sent[0][0].L = 1
#         Matrix_sent[0][0].idx1 = 0
#         Matrix_sent[0][0].idx2 = 0
#     else:
#         Matrix_sent[0][0].L = 0
#         Matrix_sent[0][0].idx1 = 0
#         Matrix_sent[0][0].idx2 = 0

#     for I in range(1, len1):
#         Matrix_sent[I][0].L = Matrix_sent[I - 1][0].L
#         Matrix_sent[I][0].idx1 = Matrix_sent[I - 1][0].idx1
#         Matrix_sent[I][0].idx2 = 0
#         if sent2[0] == sent1[I]:
#             Matrix_sent[I][0].L = 1
#             if Matrix_sent[I - 1][0].L == 0:
#                 Matrix_sent[I][0].idx1 = I

#     for I in range(1, len2):
#         Matrix_sent[0][I].L = Matrix_sent[0][I - 1].L
#         Matrix_sent[0][I].idx2 = Matrix_sent[0][I - 1].idx2
#         Matrix_sent[0][I].idx1 = 0
#         if sent2[I] == sent1[0]:
#             Matrix_sent[0][I].L = 1
#             if Matrix_sent[0][I - 1].L == 0:
#                 Matrix_sent[0][I].idx2 = I

#     for I in range(1, len1):
#         for J in range(1, len2):
#             if sent1[I] == sent2[J]:
#                 Matrix_sent[I][J].L = Matrix_sent[I - 1][J - 1].L + 1
#                 Matrix_sent[I][J].idx1 = I
#                 Matrix_sent[I][J].idx2 = J
#             elif (Matrix_sent[I - 1][J].L > Matrix_sent[I][J - 1].L):
#                 Matrix_sent[I][J].L = Matrix_sent[I - 1][J].L
#                 Matrix_sent[I][J].idx1 = Matrix_sent[I - 1][J].idx1
#                 Matrix_sent[I][J].idx2 = Matrix_sent[I - 1][J].idx2
#             else:
#                 Matrix_sent[I][J].L = Matrix_sent[I][J - 1].L
#                 Matrix_sent[I][J].idx1 = Matrix_sent[I][J - 1].idx1
#                 Matrix_sent[I][J].idx2 = Matrix_sent[I][J - 1].idx2
#     I = len1 - 1
#     J = len2 - 1
#     LP = Matrix_sent[I][J].L
#     if LP == 0:
#         return ""

#     S = ["0"] * LP
#     K = LP
#     while (I >= 0) and (J >= 0):
#         LP = Matrix_sent[I][J].L
#         if LP > 0:
#             if sent1[I] == sent2[J]:
#                 K = K - 1
#                 S[K] = sent1[I]
#                 I = I - 1
#                 J = J - 1
#             else:
#                 K1 = Matrix_sent[I][J].idx1
#                 K2 = Matrix_sent[I][J].idx2
#                 I = K1
#                 J = K2
#         else:
#             break
#     return S


def WLCS(sent1, sent2):
    len1 = len(sent1)
    len2 = len(sent2)
    if len1 == 0 or len2 == 0:
        return 0

    Matrix_sent = [[0] * len2 for i in range(len1)]
    if sent1[0] == sent2[0]:
        Matrix_sent[0][0] = 1
    else:
        Matrix_sent[0][0] = 0

    for i in range(1, len1):
        Matrix_sent[i][0] = Matrix_sent[i - 1][0]
        if sent2[0] == sent1[i]:
            Matrix_sent[i][0] = 1

    for i in range(1, len2):
        Matrix_sent[0][i] = Matrix_sent[0][i - 1]
        if sent2[i] == sent1[0]:
            Matrix_sent[0][i] = 1

    for i in range(1, len1):
        for j in range(1, len2):
            if sent1[i] == sent2[j]:
                Matrix_sent[i][j] = Matrix_sent[i - 1][j - 1] + 1
            else:
                Matrix_sent[i][j] = max(
                    Matrix_sent[i - 1][j], Matrix_sent[i][j - 1])

    i = len1 - 1
    j = len2 - 1
    LP = Matrix_sent[i][j]

    return LP

from openpyxl import load_workbook

def readTextFiles(filenames):
    textLines = []
    for filename in filenames:
        with open(filename) as f:
            lines = f.readlines()
            textLines.append(lines)
    
    return textLines




# replaceable = {'#':'Poké'}
replaceable = {}
def loadLGPEText(wb):
    result = []
    sheet = wb._sheets[0]
    # 30447
    for i in range(1,2445):
        singleItem = []
        for j in range(1,4):
            label = sheet.cell(row=i, column=j).value
            if label == None:
                label = ''
            for key in replaceable:
                label = label.replace(key,replaceable[key])
            singleItem.append(label)
        result.append(singleItem)
    return result

tmpReplaceable = {'\n':' ','!':'','?':'','\'':' ','’':' '}
jpntmpReplaceable= {'。':'','、':'','？':'','！':'','\n':'　','!':'','?':'','…':'','『':'　','@':'　@　'}
def getTextComponents(text,japanese):
    label = text
    if not japanese:
        for key in replaceable:
            label = label.replace(key,replaceable[key])
        for key in tmpReplaceable:
            label = label.replace(key,tmpReplaceable[key])
        return label.lower().split(' ')
    else:
        label = label.replace('|','\n')
        for key in jpntmpReplaceable:
            label = label.replace(key,jpntmpReplaceable[key]).replace('り','リ')
        return label.lower().split('　')





def ifIsLabel(input):
    if input == None:
        return False
    if len(input) == 0:
        return False
    if input == 'end':
        return True
    return input[-1] == ':'
    

def formatNoneText(input):
    return input or ''

def makeGen1JPNData(lines):
    result = []
    for index in range(len(lines)):
        if '-' in lines[index]:
            textLine = lines[index + 1].replace('\n','').replace('|','\n')
            tmpResult = [textLine,getTextComponents(textLine,True)]
            result.append(tmpResult)

    return result

lgpeWB = load_workbook(filename = 'xlsx/LGPE.xlsx')
loadPath = sys.argv[1]

print('Loading LGPE text')
lgpeTextList = loadLGPEText(lgpeWB)
db = lcspy.WLCSDB([getTextComponents(lgpeTextList[i][2],False) for i in range(len(lgpeTextList))])


# print('Loading Gen1 text')
# gen1TextFiles = readTextFiles(['xlsx/pkr.jp.ori.txt'])
# jpnTextdata = makeGen1JPNData(gen1TextFiles[0])

# for data in jpnTextdata:
#     print("text:")
#     print(data[0])
#     print(data[1])

# jpndb = lcspy.WLCSDB([jpnTextdata[i][1] for i in range(len(jpnTextdata))])

wb =  load_workbook(filename = loadPath)
mode = int(sys.argv[2])




for sheet in wb._sheets:
    outputPath = sheet.cell(row=1, column=1).value
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['F'].width = 10
    sheet.column_dimensions['I'].width = 10
    sheet.column_dimensions['J'].width = 34
    sheet.column_dimensions['K'].width = 34
    sheet.column_dimensions['L'].width = 34
    sheet.column_dimensions['M'].width = 34
    sheet.column_dimensions['N'].width = 40

    sheet.cell(row=1, column=mode + 6).value = '翻译区'
    sheet.cell(row=1, column=mode + 8).value = 'Gen1RB:CHS INST'
    sheet.cell(row=1, column=mode + 9).value = 'Gen1RB:CHS (Auto Match)'
    sheet.cell(row=1, column=mode + 10).value = 'Gen1RB:ENG (Auto Match)'
    sheet.cell(row=1, column=mode + 11).value = 'LGPE:ENG'
    sheet.cell(row=1, column=mode + 12).value = 'LGPE:CHS'
    sheet.cell(row=1, column=mode + 12).value = 'LGPE:JPN'
    id = 2
    while sheet.cell(row=id, column=mode).value != 'end':
        label = sheet.cell(row=id, column=mode).value
        if ifIsLabel(label):
            
            labelRow = id
            k = 1
            labelText = ''
            while not ifIsLabel(sheet.cell(row=id+k, column=mode).value):
                cell = sheet.cell(row=id+k, column=mode + 2).value
                labelText += formatNoneText(cell) + ' '
                k += 1
            
            # ind = matchWithLGPE(labelText,lgpeTextList)
            ind = db.par_search(getTextComponents(labelText,False))

            proceed = True
            lenG1 = float(len(labelText))
            lenG7 = float(len(lgpeTextList[ind][2]))

            if lenG1 < lenG7:
                proceed = lenG1 > lenG7 * 0.5
            elif lenG1 > lenG7:
                proceed = lenG7 > lenG1 * 0.5

            if True or proceed:
                perfectMatch = labelText.replace('\n','').replace(' ','') == lgpeTextList[ind][2].replace('\n','').replace(' ','')
                if perfectMatch:
                    newIDX = 0
                    tmpTXT = removeNone(sheet.cell(row=labelRow + 1 + newIDX, column=5).value)
                    while not ':' in tmpTXT and not 'end' in tmpTXT:
                        # print("HAHA " + str(labelRow + 1 + newIDX))
                        sheet.cell(row=labelRow + 1 + newIDX, column=6).value = ''
                        sheet.cell(row=labelRow + 1 + newIDX, column=7).value = ''
                        tmpTXT = removeNone(sheet.cell(row=labelRow + 1 + newIDX, column=5).value)
                        newIDX += 1
                for idx in range(len(lgpeTextList[ind])):
                    text = lgpeTextList[ind][idx]
                    lineComponents = text.split('\n')
                    



                    for lineID in range(len(lineComponents)):
                        line = lineComponents[lineID]
                        text = line.replace('\n','')
                        col = idx
                        if idx == 0:
                            col = -1
                        newcol = mode + 10 + col
                        if newcol == 12:
                            newcol = 9
                        elif newcol == 13:
                            newcol = 11
                        sheet.cell(row=labelRow + 1 + lineID, column=newcol).value = text
                        if perfectMatch:
                            sheet.cell(row=labelRow + 1 + lineID, column=newcol).fill = warningFill
                        if perfectMatch:
                            if newcol == 9:
                                newcol = 6
                            elif newcol == 10:
                                newcol = 7
                        sheet.cell(row=labelRow + 1 + lineID, column=newcol).value = text
                        if perfectMatch:
                            sheet.cell(row=labelRow + 1 + lineID, column=newcol).fill = warningFill
                        # if proceed:
                        #     sheet.cell(row=labelRow + 1 + lineID, column=newcol).fill = warningFill

                # lgpeJPNTextComponent = getTextComponents(lgpeTextList[ind][1],True)
                # ind2 =jpndb.par_search(lgpeJPNTextComponent)


                # text = jpnTextdata[ind2][0]
                # lineComponents = text.split('\n')
                # for lineID in range(len(lineComponents)):
                #     line = lineComponents[lineID]
                #     text = line.replace('\n','')
                #     sheet.cell(row=labelRow + 1 + lineID, column=mode + 10).value = text
                #     if proceed:
                #         sheet.cell(row=labelRow + 1 + lineID, column=mode + 10).fill = warningFill
                # print(labelText.replace('\n','').replace(' ','') == lgpeTextList[ind][2].replace('\n','').replace(' ',''))
                print('Original Text:')
                print(labelText.replace('\n','').replace(' ',''))
                print("Most Match LGPE ENG:")
                print(lgpeTextList[ind][2].replace('\n','').replace(' ',''))
                # print("Most Match LGPE CHS:")
                # print(lgpeTextList[ind][0])
                # print("Most Match LGPE JPN:")
                # print(lgpeTextList[ind][1])
                # print("Most Match Gen1 JPN:")
                # print(jpnTextdata[ind2][0])

                print('')
            
        id += 1

wb.save(loadPath)