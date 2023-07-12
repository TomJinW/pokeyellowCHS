from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border

import sys
import openpyxl
import shutil
import os
import charmap

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

def addFilePaths(path):
    if not path in filePaths:
        filePaths.append(path)
        os.makedirs(os.path.dirname(tmp+path), exist_ok=True)
        # shutil.copyfile(path,tmp+path)

def readTextFile(path):
    file = open(path,"r")
    return file.read()

def readTextFileLines(path):
    file = open(path,"r")
    return file.readlines()

def removeNone(text):
    if text == None:
        return ''
    return text

tmp = 'tmp/'
# Program Start
xlsxListsPath = sys.argv[1]
relativePath = sys.argv[2]
mode = int(sys.argv[3])
print(bcolors.OKCYAN)
workbookPaths = readTextFileLines(xlsxListsPath)
filePaths = []

for workbookPath in workbookPaths:
    # Load Workbook
    newPath =  relativePath + workbookPath.replace('\n','') 
    print('opening ' + newPath)
    wb = load_workbook(filename = newPath)
    charMap = charmap.readCharMaps()
    for sheet in wb._sheets:
        # Get init File Path
        filePath = sheet.cell(row=1, column=1).value
        # Backup Original Files
        addFilePaths(filePath)
        id = 2
        while sheet.cell(row=id, column=1).value != 'end':
            if sheet.cell(row=id, column=1).value != None:
                if sheet.cell(row=id, column=1).value != 'end' and not':' in sheet.cell(row=id, column=1).value:
                    filePath = sheet.cell(row=id, column=1).value
                    addFilePaths(filePath)
            id += 1


if mode != 0:
    print('Restoring...')
for path in filePaths:
    if mode == 0:
        print('Backing up: ' + path)
        try:
            shutil.copyfile(path,tmp+path)
        except:
            print('caught something that is not a file!')
            print(path)
       
    else:
        # print('Restoring: ' + path)
        try:
            shutil.copyfile(tmp+path,path)
        except:
            print('caught something that is not a file!')
            print(path)

print('Backup / Restore Complete!')