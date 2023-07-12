from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border

import sys
import openpyxl
import shutil
import os
import charmap

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

def addFilePaths(path):
    if not path in filePaths:
        filePaths.append(path)
        os.makedirs(os.path.dirname(tmp+path), exist_ok=True)
        

def readTextFile(path):
    file = open(path,"r")
    return file.read()

def removeNone(text):
    if text == None:
        return ''
    return text

def replaceTextwithCondition(text,replacee,replacer,lastRow):
    
    lines = text.split('\n')
    output = ''
    if len(lines) > 0:
        output = lines[0] +' \n'
    for i in range(1,len(lines)):
        line = lines[i]
        lastLine = lines[i - 1]
        if lastRow in lastLine:
            if buildMode != 2:
                output += line.replace(replacee,replacer + ' from: ' + replacee) + '\n'
            else:
                output += line.replace(replacee,replacer) + '\n'
        else:
            output += line  + '\n'
    return output
# def replaceTextwithCondition(origText,text,replacee,replacer,lastRow):
    
#     lines = text.split('\n')
#     origLines = origText.split('\n')
#     output = ''
#     if len(lines) > 0:
#         output = lines[0] +' \n'
#     for i in range(1,len(lines)):
#         line = lines[i]
#         lastLine = origLines[i - 1]
#         if lastRow in lastLine:
#             output += line.replace(replacee,replacer + ' from: ' + replacee) + '\n'
#         else:
#             output += line  + '\n'
#     return output

tmp = 'tmp/'
# Program Start
xlsxListPath = sys.argv[1]
mode = int(sys.argv[2])
ver = sys.argv[3]
buildMode = int(sys.argv[4])
# Load Workbook
wb = load_workbook(filename = xlsxListPath)

filePaths = []

charMap = charmap.readCharMaps()

print()
print(bcolors.OKGREEN)
print('Importing db Data...')

def getIfSkipped(inputVer):
    if inputVer == '':
        return False
    return inputVer != ver

for sheet in wb._sheets:
    replacees = []
	# Get init File Path
    filePath = sheet.cell(row=1, column=mode).value
    tmpVer = removeNone(sheet.cell(row=1, column=mode + 2).value)
    skipped = getIfSkipped(tmpVer)
    endChar = removeNone(sheet.cell(row=1, column=mode+1).value)
	# Backup Original Files
    addFilePaths(filePath)
    text2Modify = readTextFile(filePath)
    originalText = readTextFile(filePath)

    id = 2
    while sheet.cell(row=id, column=mode).value != 'end':

        if sheet.cell(row=id, column=mode).value != None:
            if sheet.cell(row=id, column=mode).value != 'end':
                with open(filePath, 'w') as f:
                    f.write(text2Modify)
                replacees = []
                filePath = sheet.cell(row=id, column=mode).value
                tmpVer = removeNone(sheet.cell(row=id, column=mode + 2).value)
                skipped = getIfSkipped(tmpVer)
                endChar = removeNone(sheet.cell(row=id, column=mode+1).value)
                addFilePaths(filePath)
                text2Modify = readTextFile(filePath)
                id += 1
                continue

        replacee = removeNone(sheet.cell(row=id, column = mode + 1).value)
        replacer = removeNone(sheet.cell(row=id, column = mode + 2).value) + endChar

        if replacee != '' and (not skipped):
            repetitive = False
            for currentsReplacee in replacees:
                if replacee in currentsReplacee:
                    repetitive = True
                    print(bcolors.OKBLUE + "Warning! Duplicate: " + replacee)
            if not repetitive:
                replacees.append(replacee)    

            if sheet.cell(row=id, column = mode + 4).value == None:
                if sheet.cell(row=id, column = mode + 3).value != None:
                    newReplacee = sheet.cell(row=id, column = mode + 3).value
                    text2Modify = text2Modify.replace(replacee,newReplacee)
                else:
                    if buildMode != 2:
                        text2Modify = text2Modify.replace(replacee,charmap.replaceText(replacer,charMap,buildMode) + ' from: ' + replacee)
                    else:
                        text2Modify = text2Modify.replace(replacee,charmap.replaceText(replacer,charMap,buildMode))
            else:
                lastRow = sheet.cell(row=id, column = mode + 4).value
                if sheet.cell(row=id, column = mode + 3).value != None:
                    newReplacee = sheet.cell(row=id, column = mode + 3).value
                    text2Modify = replaceTextwithCondition(text2Modify,replacee,newReplacee,lastRow)
                else:
                    text2Modify = replaceTextwithCondition(text2Modify,replacee,charmap.replaceText(replacer,charMap,buildMode),lastRow)
        id += 1

    # print(text2Modify)
    with open(filePath, 'w') as f:
        f.write(text2Modify)
# input(bcolors.OKGREEN + "Any key..")
print(bcolors.OKGREEN)
print()
print('db Data Import complete.')
input("Press Return to proceed..")



