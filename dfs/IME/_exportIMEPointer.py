from openpyxl import load_workbook
# import sys
# from openpyxl.styles import Color, PatternFill, Font, Border
# import openpyxl

def addAtSymbol(input):
    output = input
    length = len(input)
    for i in range(7-length):
        output += '@'
    return output

wb = load_workbook(filename = 'char.xlsx')

dicts =[{},{},{},{},{},{}] 
for sheet in wb._sheets:
    for row in range(1,6728):
        for col in range(2,5):
            pin = sheet.cell(row=row, column=col).value
            if pin != None:
                length = len(pin)
                if not pin in dicts[length - 1]:
                    dicts[length - 1][pin] = 1
                else:
                    dicts[length - 1][pin] += 1

print('PinyinStrTableInfoTable:')
for i in range(0,6):
    dict = dicts[i]
    print('\tdb ' + str(len(dict)))

for i in range(0,6):
    dict = dicts[i]
    print('PinyinStrTableLen'+ str(i + 1) + ':')
    for key in dict:
        print('\tdb \"'+ addAtSymbol(key).upper() + '\"')
        print('\tdw IME_'+ key.upper() + '_Char')
        count = dict[key]
        pagenum = int(count / 12)
        if count % 12 != 0:
            pagenum += 1
        print('\tdb '+ str(pagenum))
        print('\t;'+ str(count))
    print()