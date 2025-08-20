from openpyxl import load_workbook
# import sys
# from openpyxl.styles import Color, PatternFill, Font, Border
# import openpyxl

def addAtSymbol(input):
    output = input
    length = len(input)
    for i in range(7-length):
        output += '@'
    return output

def modHexText(input):
    output = input
    output = output.replace('HEX','$')
    outputlist = list(output)
    outputlist.insert(3, ',')
    outputlist.insert(4, '$')
    output = ''.join(outputlist)
    output += ','
    return output

wb = load_workbook(filename = 'char.xlsx')

codeTable = {}
for sheet in wb._sheets:
    for row in range(1,6769):
        char = sheet.cell(row=row, column=6).value
        hexStr = modHexText(sheet.cell(row=row, column=7).value)
        codeTable[char] = hexStr

charTable = {}
for sheet in wb._sheets:
    for row in range(1,6728):
        for col in range(2,5):
            pin = sheet.cell(row=row, column=col).value
            char = sheet.cell(row=row, column=1).value
            if pin != None:
                if not pin in charTable:
                    charTable[pin] = [char]
                else:
                    charTable[pin].append(char)

# print(charTable)
for key in charTable:
    # if key.upper() == '***':
    #     print(key)
    print('IME_'+key.upper()+'_Char:')
    text = '\t db '
    chs = '\t ; '
    for i in range(len(charTable[key])):
        item = charTable[key][i]
        code = codeTable[item]
        text += code
        chs += item + ' '
    text += '$50'
    
    print(text)
    print(chs)

print('\t db $50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50')
